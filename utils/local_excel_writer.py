"""
로컬 엑셀 파일 직접 쓰기 모듈 (OneDrive 동기화 활용)

Graph API의 Workbook 엔드포인트가 66시트 복잡 구조 때문에 501 에러를 내므로,
대신 로컬 OneDrive 동기화 폴더의 엑셀 파일을 openpyxl로 직접 수정한다.
OneDrive가 자동으로 SharePoint에 동기화해서 다른 사용자(모바일 조회 등)에게도 반영.

작동 조건:
- 세웅 PC(집/회사) 또는 실무자 PC에서 Streamlit 실행 시
- 해당 PC에 OneDrive 동기화가 설정되어 있어야 함
- 대상 엑셀 파일이 Excel로 열려있지 않아야 함 (쓰기 충돌 방지)
"""
import os
from datetime import datetime
from openpyxl import load_workbook


# ── 로컬 엑셀 파일 경로 후보 ──
LOCAL_EXCEL_CANDIDATES = [
    # 집PC
    r"C:\Users\Hwang\OneDrive - GS칼텍스 예울마루\DAX\yeulmaru_dashboard_db_v1.xlsx",
    # 회사PC (한글 사용자명)
    r"C:\Users\황세웅\OneDrive - GS칼텍스 예울마루\DAX\yeulmaru_dashboard_db_v1.xlsx",
]


def find_local_excel_path():
    """운영 엑셀 파일의 로컬 경로 자동 탐지.

    Returns:
        str | None: 존재하는 첫 번째 경로, 없으면 None
    """
    userprofile = os.environ.get('USERPROFILE', '')
    if userprofile:
        dynamic = os.path.join(
            userprofile,
            "OneDrive - GS칼텍스 예울마루",
            "DAX",
            "yeulmaru_dashboard_db_v1.xlsx",
        )
        candidates = [dynamic] + LOCAL_EXCEL_CANDIDATES
    else:
        candidates = LOCAL_EXCEL_CANDIDATES

    for path in candidates:
        if os.path.exists(path):
            return path
    return None


def _find_cumulative_header_row(ws, max_scan=30):
    """'일일입력' 시트에서 누적기록 헤더 row 찾기.

    조건: A열='기준일자' AND B열='공연명' 인 row.
    (row 2의 '기준일자' 레이블과 구분하기 위함)
    """
    for r in range(1, max_scan + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        if a == '기준일자' and b == '공연명':
            return r
    return None


def _find_last_data_row(ws, header_row):
    """누적기록 헤더 아래에서 마지막 리터럴 yyyymmdd 데이터 row (위→아래 탐색)."""
    last = header_row
    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a is None:
            continue
        try:
            a_int = int(float(a))
            if 20000000 < a_int < 30000000:
                last = r
        except (ValueError, TypeError):
            continue
    return last


def _find_insert_row(ws, header_row):
    """INSERT 대상 row 결정.

    A파일 구조: 누적 데이터 뒤에 안내 텍스트/수식 '섬' 존재 (row 2653~2663).
    이를 건너뛰고 그 뒤 빈 영역 첫 row 반환.

    로직:
      1. 마지막 yyyymmdd 데이터 row (last_data) 찾기
      2. last_data + 1 부터 스캔:
         - A열이 None → 빈 row, 여기 반환
         - A열이 yyyymmdd → 또 다른 데이터 row (이전 INSERT 결과), 계속
         - A열이 그 외 (안내 텍스트 등) → 건너뜀, 계속
      3. 끝까지 못 찾으면 max_row + 1 반환
    """
    last_data = _find_last_data_row(ws, header_row)

    for r in range(last_data + 1, ws.max_row + 2):
        a = ws.cell(row=r, column=1).value
        if a is None:
            return r
        try:
            a_int = int(float(a))
            if 20000000 < a_int < 30000000:
                continue
        except (ValueError, TypeError):
            pass
    return ws.max_row + 1


def _find_matching_row(ws, header_row, date_int, perf_name, round_time):
    """기준일자+공연명+회차시각 매칭 행 탐색.

    Returns:
        int | None: 매칭된 row (1-based), 없으면 None
    """
    perf_s = str(perf_name).strip()
    rt_s = str(round_time).strip().lstrip("'") if round_time else ""

    for r in range(header_row + 1, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if a is None:
            continue
        try:
            a_int = int(float(a))
        except (ValueError, TypeError):
            continue
        if a_int != date_int:
            continue

        b = ws.cell(row=r, column=2).value
        b_str = str(b).strip() if b else ""
        if not (b_str == perf_s or perf_s in b_str or b_str in perf_s):
            continue

        d = ws.cell(row=r, column=4).value
        d_str = str(d).strip().lstrip("'") if d else ""
        if rt_s and d_str and rt_s != d_str:
            continue

        return r
    return None



def save_daily_entry_local(
    date_int, perf_name, perf_date_str, round_time,
    open_seats, paid_seats, paid_amount,
    rsv_seats, rsv_amount, free_seats,
    prev_seats=0, prev_amount=0,
):
    """일일입력 시트 누적기록에 1행 저장 (로컬 파일 직접 쓰기).

    Graph API Workbook 엔드포인트 대신 로컬 OneDrive 동기화 파일을 수정.
    OneDrive가 자동으로 SharePoint와 동기화.

    Returns:
        dict: {status: 'updated'|'inserted'|'error', row: int|None, message: str}
    """
    # 1. 계산
    total_seats = paid_seats + rsv_seats + free_seats
    total_amount = paid_amount + rsv_amount
    occupancy = min(total_seats / open_seats, 1.0) if open_seats > 0 else 0
    unit_price = round(total_amount / total_seats) if total_seats > 0 else 0
    diff_seats = total_seats - prev_seats
    diff_amount = total_amount - prev_amount
    now_time = datetime.now().strftime('%H:%M:%S')

    # 2. 로컬 파일 탐지
    local_path = find_local_excel_path()
    if not local_path:
        return {
            "status": "error",
            "row": None,
            "message": (
                "로컬 엑셀 파일을 찾을 수 없습니다. "
                "OneDrive 동기화가 설정된 PC에서 실행해야 저장 가능합니다. "
                f"확인 경로: {LOCAL_EXCEL_CANDIDATES}"
            ),
        }

    # 3. 데이터 행 구성 (A~R, 18컬럼)
    row_data = [
        date_int,                                      # A: 기준일자
        perf_name,                                     # B: 공연명
        perf_date_str,                                 # C: 공연일
        ("'" + round_time) if round_time else '',      # D: 회차/시각
        open_seats,                                    # E: 오픈석
        paid_seats,                                    # F: 유료좌석
        paid_amount,                                   # G: 유료금액
        rsv_seats,                                     # H: 예약좌석
        rsv_amount,                                    # I: 예약금액
        free_seats,                                    # J: 무료좌석
        total_seats,                                   # K: 합계좌석
        total_amount,                                  # L: 합계금액
        round(occupancy, 4),                           # M: 점유율
        diff_seats,                                    # N: 전일대비(석)
        diff_amount,                                   # O: 전일대비(원)
        unit_price,                                    # P: 객단가
        "",                                            # Q: 중복체크
        now_time,                                      # R: 갱신시각
    ]

    # 4. 파일 열기
    try:
        wb = load_workbook(local_path)
    except PermissionError as e:
        return {
            "status": "error",
            "row": None,
            "message": f"파일 열기 실패 (Excel로 열려있는지 확인): {e}",
        }
    except Exception as e:
        return {"status": "error", "row": None, "message": f"파일 로드 실패: {e}"}

    try:
        if '일일입력' not in wb.sheetnames:
            wb.close()
            return {"status": "error", "row": None,
                    "message": "'일일입력' 시트를 찾을 수 없습니다."}
        ws = wb['일일입력']

        header_row = _find_cumulative_header_row(ws)
        if header_row is None:
            wb.close()
            return {"status": "error", "row": None,
                    "message": "누적기록 헤더 행을 찾을 수 없습니다."}

        # 5. 매칭 행 탐색
        existing_row = _find_matching_row(ws, header_row, date_int, perf_name, round_time)

        if existing_row:
            # UPDATE
            target_row = existing_row
            status = "updated"
        else:
            # INSERT: 안내 텍스트 섬 건너뛴 빈 영역 첫 행
            target_row = _find_insert_row(ws, header_row)
            status = "inserted"

        # 6. 행에 데이터 쓰기
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=target_row, column=col_idx).value = val

        # 7. 저장
        wb.save(local_path)
        wb.close()

        return {
            "status": status,
            "row": target_row,
            "message": (
                f"행 {target_row} 갱신 완료 ({now_time})"
                if status == "updated"
                else f"행 {target_row} 신규 저장 ({now_time})"
            ),
        }
    except PermissionError as e:
        try:
            wb.close()
        except Exception:
            pass
        return {
            "status": "error",
            "row": None,
            "message": f"파일 저장 실패 (Excel로 열려있는지 확인): {e}",
        }
    except Exception as e:
        try:
            wb.close()
        except Exception:
            pass
        return {"status": "error", "row": None, "message": f"저장 중 예외: {e}"}
